import pandas as pd
from sqlalchemy import create_engine, exc
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

def get_engine_local():

    try:
        engine = create_engine("postgresql+psycopg2://{user}:{pw}@{host}:{port}/{dbname}".format(
                user = "postgres",
                pw = "Brutal.shred01",
                host = "localhost",
                dbname = "chinook",
                port = "5432"
            ))
        return engine
    except Exception as e:
        print(f"Hiba az adatkapcsolat létrehozásakot: {e}")

def main():

    engine = get_engine_local()

    query = """
    SELECT 
    CONCAT(c."first_name", ' ', c."last_name") AS fullname,
    c. "customer_id",
    SUM(i.total) AS total_spent
    FROM customer AS c
    INNER JOIN invoice i ON c.customer_id = i.customer_id
    GROUP BY c.customer_id, fullname
    HAVING SUM(i.total) >= 50
    ORDER BY total_spent DESC;
    """

    try:
        with engine.connect() as connection:
            df = pd.read_sql(query, connection)
            excel_path = 'C:/Users/diefi/Documents/GitHub/Python_learn/Exercise_topics/reports/report.xlsx'
            df.to_excel(excel_path, index=False, engine="openpyxl")

        format_excel(excel_path)
        print("A report sikeresen elékszült.")
    except Exception as e:
        print(f"Hiba a report elkészítésekor: {e}")

def format_excel(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="AAD8E6", end_color="AAD8E6", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    wb.save(excel_path)
    wb.close()


if __name__ == '__main__':
    main()