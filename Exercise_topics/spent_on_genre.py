import pandas as pd
from sqlalchemy import create_engine, exc
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime

def get_engine_local():

    try:
        engine = create_engine("postgresql+psycopg2://{user}:{pw}@{host}:{port}/{dbname}".format(
                user = "postgres",
                pw = "Brutal.shred01",
                host = "localhost",
                dbname = "chinook",
                port = "5432"
            ))
        return engine
    except Exception as e:
        print(f"Hiba az adatkapcsolat létrehozásakot: {e}")

def main():

    engine = get_engine_local()

    query = """
    WITH RankedGenres AS (
	SELECT
	c.customer_id,
	CONCAT(c.first_name, ' ', c.last_name) AS fullname,
	SUM(i.total) AS total_spent,
	g.name AS genre,
	RANK() OVER (PARTITION BY c.customer_id ORDER BY SUM(i.total) DESC) AS genre_rank 
	FROM customer c
	INNER JOIN invoice i ON c.customer_id = i.customer_id
	INNER JOIN invoice_line il ON i.invoice_id = il.invoice_id
	INNER JOIN track t ON il.track_id = t.track_id
	INNER JOIN genre g ON t.genre_id = g.genre_id
	GROUP BY c.customer_id, fullname, g.name
    )
    SELECT customer_id, fullname, total_spent, genre AS top_genre
    FROM RankedGenres 
    WHERE genre_rank = 1
    ORDER BY total_spent DESC
    LIMIT 5;
    """

    try:
        with engine.connect() as connection:
            df = pd.read_sql(query, connection)
            excel_path = 'C:/Users/diefi/Documents/GitHub/Python_learn/Exercise_topics/reports/top_spenders.xlsx'
            df.to_excel(excel_path, index=False, engine="openpyxl")

        format_excel(excel_path)
        print("A report sikeresen elékszült.")
    except Exception as e:
        print(f"Hiba a report elkészítésekor: {e}")

def format_excel(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="008040", end_color="008040", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    last_row = ws.max_row + 1
    ws[f"A{last_row}"] = f"Report generated: {timestamp}"
    ws[f"A{last_row}"].font = Font(italic=True, color="808080")

    wb.save(excel_path)
    wb.close()


if __name__ == '__main__':
    main()